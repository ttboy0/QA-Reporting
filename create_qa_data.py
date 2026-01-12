"""
Script to create the QA Data Excel template
Run this once to generate qa_data.xlsx
"""

import openpyxl
import os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create a new workbook
wb = openpyxl.Workbook()

# Remove default sheet and create new ones
wb.remove(wb.active)
ws_api = wb.create_sheet("API Data", 0)
ws_web = wb.create_sheet("Web Data", 1)

# Define styles
header_fill = PatternFill(start_color="2C2C54", end_color="2C2C54", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
subheader_fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
subheader_font = Font(bold=True, color="2C2C54", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal="center", vertical="center")

# ============= API DATA SHEET =============
ws_api['A1'] = "API TESTING STATUS REPORT - DATA ENTRY"
ws_api['A1'].font = Font(bold=True, size=14, color="2C2C54")
ws_api.merge_cells('A1:F1')

# Report Metadata
ws_api['A3'] = "Report Period:"
ws_api['B3'] = "Week of Jan 6-12, 2026"
ws_api['A4'] = "API Test Lead:"
ws_api['B4'] = "David Park"
ws_api['A5'] = "Lead Email:"
ws_api['B5'] = "david.park@company.com"

# API Summary Table
ws_api['A7'] = "API TEST SUITES SUMMARY"
ws_api['A7'].font = subheader_font
ws_api['A7'].fill = subheader_fill
ws_api.merge_cells('A7:G7')

headers = ["Test Suite", "Total Tests", "Passed", "Failed", "Blocked", "Pass Rate %", "Status"]
for col, header in enumerate(headers, 1):
    cell = ws_api.cell(row=8, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border

# API Test Suites Data
api_suites = [
    ["Authentication API", 156, 152, 3, 1, 97, "Stable"],
    ["Payment Processing", 289, 268, 18, 3, 92, "Monitor"],
    ["Inventory Management", 198, 187, 9, 2, 94, "Stable"],
    ["Reporting Engine", 204, 182, 18, 4, 89, "Monitor"],
]

for row_idx, suite in enumerate(api_suites, 9):
    for col_idx, value in enumerate(suite, 1):
        cell = ws_api.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border
        if col_idx in [2, 3, 4, 5, 6]:  # Numeric columns
            cell.alignment = center_align

# API Totals Row
ws_api['A13'] = "TOTALS"
ws_api['A13'].font = Font(bold=True)
ws_api['B13'] = "=SUM(B9:B12)"
ws_api['C13'] = "=SUM(C9:C12)"
ws_api['D13'] = "=SUM(D9:D12)"
ws_api['E13'] = "=SUM(E9:E12)"
ws_api['F13'] = "=ROUND(C13/B13*100,0)"
for col in range(1, 7):
    ws_api.cell(row=13, column=col).font = Font(bold=True)
    ws_api.cell(row=13, column=col).border = border

# Defect by Priority
ws_api['A15'] = "DEFECT BREAKDOWN BY PRIORITY"
ws_api['A15'].font = subheader_font
ws_api['A15'].fill = subheader_fill
ws_api.merge_cells('A15:C15')

defect_headers = ["Priority", "Count", "Status"]
for col, header in enumerate(defect_headers, 1):
    cell = ws_api.cell(row=16, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border

defect_data = [
    ["Critical", 2, "Resolved"],
    ["High", 8, "In Progress"],
    ["Medium", 22, "Scheduled"],
    ["Low", 16, "Backlog"],
]

for row_idx, defect in enumerate(defect_data, 17):
    for col_idx, value in enumerate(defect, 1):
        cell = ws_api.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border

# Coverage Percentage
ws_api['A22'] = "AUTOMATION COVERAGE BY AREA"
ws_api['A22'].font = subheader_font
ws_api['A22'].fill = subheader_fill
ws_api.merge_cells('A22:B22')

coverage_headers = ["Area", "Coverage %"]
for col, header in enumerate(coverage_headers, 1):
    cell = ws_api.cell(row=23, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border

coverage_data = [
    ["Authentication", 90],
    ["Payment", 85],
    ["Inventory", 60],
    ["Reporting", 75],
]

for row_idx, cov in enumerate(coverage_data, 24):
    for col_idx, value in enumerate(cov, 1):
        cell = ws_api.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border
        if col_idx == 2:
            cell.alignment = center_align

# Risks Section
ws_api['A29'] = "RISKS & HIGH PRIORITY ISSUES"
ws_api['A29'].font = subheader_font
ws_api['A29'].fill = subheader_fill
ws_api.merge_cells('A29:E29')

risk_headers = ["Issue ID", "Description", "Priority", "Assigned Owner", "Target Date"]
for col, header in enumerate(risk_headers, 1):
    cell = ws_api.cell(row=30, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border

risk_data = [
    ["API-001", "Payment Gateway: Multi-currency edge cases failing", "HIGH", "Michael Chen", "Jan 15"],
    ["API-002", "Reporting Engine: Performance degradation >500 req/s", "HIGH", "David Park", "Jan 16"],
    ["API-003", "Authentication: Session timeout edge case", "MEDIUM", "Michael Chen", "Jan 11"],
]

for row_idx, risk in enumerate(risk_data, 31):
    for col_idx, value in enumerate(risk, 1):
        cell = ws_api.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border

# Set column widths
ws_api.column_dimensions['A'].width = 25
ws_api.column_dimensions['B'].width = 15
ws_api.column_dimensions['C'].width = 12
ws_api.column_dimensions['D'].width = 12
ws_api.column_dimensions['E'].width = 12
ws_api.column_dimensions['F'].width = 12
ws_api.column_dimensions['G'].width = 15

# ============= WEB DATA SHEET =============
ws_web['A1'] = "WEB TESTING STATUS REPORT - DATA ENTRY"
ws_web['A1'].font = Font(bold=True, size=14, color="2C2C54")
ws_web.merge_cells('A1:F1')

# Report Metadata
ws_web['A3'] = "Report Period:"
ws_web['B3'] = "Week of Jan 6-12, 2026"
ws_web['A4'] = "UI Test Lead:"
ws_web['B4'] = "Jessica Martinez"
ws_web['A5'] = "Lead Email:"
ws_web['B5'] = "jessica.martinez@company.com"

# Web Summary Table
ws_web['A7'] = "WEB TEST SUITES SUMMARY"
ws_web['A7'].font = subheader_font
ws_web['A7'].fill = subheader_fill
ws_web.merge_cells('A7:G7')

for col, header in enumerate(headers, 1):
    cell = ws_web.cell(row=8, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border

# Web Test Suites Data
web_suites = [
    ["Login & Auth Flow", 78, 76, 2, 0, 97, "Stable"],
    ["Checkout Flow", 92, 78, 12, 2, 85, "Monitor"],
    ["Product Search", 68, 61, 5, 2, 90, "Good"],
    ["Dashboard & Reports", 82, 65, 14, 3, 79, "Action"],
]

for row_idx, suite in enumerate(web_suites, 9):
    for col_idx, value in enumerate(suite, 1):
        cell = ws_web.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border
        if col_idx in [2, 3, 4, 5, 6]:
            cell.alignment = center_align

# Web Totals Row
ws_web['A13'] = "TOTALS"
ws_web['A13'].font = Font(bold=True)
ws_web['B13'] = "=SUM(B9:B12)"
ws_web['C13'] = "=SUM(C9:C12)"
ws_web['D13'] = "=SUM(D9:D12)"
ws_web['E13'] = "=SUM(E9:E12)"
ws_web['F13'] = "=ROUND(C13/B13*100,0)"
for col in range(1, 7):
    ws_web.cell(row=13, column=col).font = Font(bold=True)
    ws_web.cell(row=13, column=col).border = border

# Defect by Priority
ws_web['A15'] = "DEFECT BREAKDOWN BY PRIORITY"
ws_web['A15'].font = subheader_font
ws_web['A15'].fill = subheader_fill
ws_web.merge_cells('A15:C15')

for col, header in enumerate(defect_headers, 1):
    cell = ws_web.cell(row=16, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border

web_defect_data = [
    ["Critical", 2, "Resolved"],
    ["High", 4, "In Progress"],
    ["Medium", 10, "Scheduled"],
    ["Low", 5, "Backlog"],
]

for row_idx, defect in enumerate(web_defect_data, 17):
    for col_idx, value in enumerate(defect, 1):
        cell = ws_web.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border

# Coverage Percentage
ws_web['A22'] = "AUTOMATION COVERAGE BY AREA"
ws_web['A22'].font = subheader_font
ws_web['A22'].fill = subheader_fill
ws_web.merge_cells('A22:B22')

for col, header in enumerate(coverage_headers, 1):
    cell = ws_web.cell(row=23, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border

web_coverage_data = [
    ["Login & Auth", 99],
    ["Checkout Flow", 77],
    ["Product Search", 44],
    ["Dashboard", 55],
]

for row_idx, cov in enumerate(web_coverage_data, 24):
    for col_idx, value in enumerate(cov, 1):
        cell = ws_web.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border
        if col_idx == 2:
            cell.alignment = center_align

# Risks Section
ws_web['A29'] = "RISKS & HIGH PRIORITY ISSUES"
ws_web['A29'].font = subheader_font
ws_web['A29'].fill = subheader_fill
ws_web.merge_cells('A29:E29')

for col, header in enumerate(risk_headers, 1):
    cell = ws_web.cell(row=30, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border

web_risk_data = [
    ["SEL-001", "Checkout Flow: UI elements not rendering on Safari", "HIGH", "Jessica Martinez", "Jan 14"],
    ["SEL-002", "Dashboard: Data table pagination failing in Firefox", "MEDIUM", "Robert Thompson", "Jan 20"],
    ["SEL-003", "Mobile Responsive: Button alignment issue on Android", "MEDIUM", "Jessica Martinez", "Jan 22"],
]

for row_idx, risk in enumerate(web_risk_data, 31):
    for col_idx, value in enumerate(risk, 1):
        cell = ws_web.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border

# Set column widths
ws_web.column_dimensions['A'].width = 25
ws_web.column_dimensions['B'].width = 15
ws_web.column_dimensions['C'].width = 12
ws_web.column_dimensions['D'].width = 12
ws_web.column_dimensions['E'].width = 12
ws_web.column_dimensions['F'].width = 12
ws_web.column_dimensions['G'].width = 15

# Save the workbook
if not os.path.exists('qa_data.xlsx'):
    wb.save('qa_data.xlsx')
    print("✓ qa_data.xlsx created successfully!")
    print("  - Open this file to edit your QA data")
else:
    print("✓ qa_data.xlsx already exists. Skipping creation.")

