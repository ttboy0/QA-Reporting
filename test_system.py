"""
Comprehensive Test Suite for QA Reporting System
Tests report generation, data integrity, chart accuracy, and output validation
"""

import os
import shutil
import json
import re
import pandas as pd
from pathlib import Path
from generate_all import generate_html_report, export_html, read_excel_data

class TestSuite:
    def __init__(self):
        self.test_results = []
        self.excel_file = 'qa_data.xlsx'
        self.reports_dir = 'reports'
        self.backup_file = 'qa_data_backup.xlsx'
        
    def log_test(self, test_name, passed, message=""):
        """Log test result"""
        status = "✓ PASS" if passed else "✗ FAIL"
        print(f"{status}: {test_name}")
        if message:
            print(f"       {message}")
        self.test_results.append({'test': test_name, 'passed': passed, 'message': message})
        
    def test_1_report_generation(self):
        """Test 1: Verify HTML reports are generated"""
        print("\n[TEST 1] Report Generation")
        print("-" * 60)
        
        # Clean reports directory
        if os.path.exists(self.reports_dir):
            shutil.rmtree(self.reports_dir)
        
        # Generate reports
        try:
            for report_type in ["api", "web"]:
                html_content, _ = generate_html_report(report_type)
                export_html(html_content, report_type)
            
            # Check files exist
            api_exists = os.path.exists(f"{self.reports_dir}/api_report.html")
            web_exists = os.path.exists(f"{self.reports_dir}/web_report.html")
            
            self.log_test("API report generated", api_exists)
            self.log_test("Web report generated", web_exists)
            
            # Check no PDF or PPTX files
            pdf_exists = any(f.endswith('.pdf') for f in os.listdir(self.reports_dir))
            pptx_exists = any(f.endswith('.pptx') for f in os.listdir(self.reports_dir))
            
            self.log_test("No PDF files generated", not pdf_exists, 
                         "PDF export is not supported" if pdf_exists else "")
            self.log_test("No PPTX files generated", not pptx_exists,
                         "PPTX export is not supported" if pptx_exists else "")
            
            return api_exists and web_exists and not pdf_exists and not pptx_exists
            
        except Exception as e:
            self.log_test("Report generation", False, str(e))
            return False
    
    def test_2_api_data_modifications(self):
        """Test 2: Verify API data modifications are reflected in reports"""
        print("\n[TEST 2] API Data Modifications")
        print("-" * 60)
        
        try:
            # Backup original file
            shutil.copy(self.excel_file, self.backup_file)
            
            from openpyxl import load_workbook
            wb_modify = load_workbook(self.excel_file)
            ws_api = wb_modify['API Data']
            
            # Modify Test Suite (Row 9 in openpyxl)
            ws_api['A9'] = "MODIFIED_API_SUITE_XYZ"
            ws_api['B9'] = 999
            ws_api['C9'] = 888
            ws_api['D9'] = 77
            
            # Modify Defect Breakdown (Row 17-18 in openpyxl)
            ws_api['B17'] = 111  # Critical
            ws_api['B18'] = 222  # High
            
            # Modify Coverage (Row 23-24 in openpyxl)
            ws_api['B23'] = 98   # Authentication
            ws_api['B24'] = 96   # Payment
            
            # Modify Risks (Row 31 in openpyxl)
            ws_api['B31'] = "MODIFIED_API_RISK_DESCRIPTION_XYZ"
            ws_api['D31'] = "MODIFIED_API_OWNER_XYZ"
            
            wb_modify.save(self.excel_file)
            
            # Generate new report
            html_api, _ = generate_html_report("api")
            
            # --- Assertions ---
            self.log_test("API Suite Name Change", "MODIFIED_API_SUITE_XYZ" in html_api)
            self.log_test("API Total Tests Change", "999" in html_api)
            self.log_test("API Passed Tests Change", "888" in html_api)
            self.log_test("API Failed Tests Change", "77" in html_api)
            self.log_test("API Critical Defects Change", "111" in html_api)
            self.log_test("API High Defects Change", "222" in html_api)
            self.log_test("API Authentication Coverage Change", "98" in html_api)
            self.log_test("API Payment Coverage Change", "96" in html_api)
            self.log_test("API Risk Description Change", "MODIFIED_API_RISK_DESCRIPTION_XYZ" in html_api)
            self.log_test("API Risk Owner Change", "MODIFIED_API_OWNER_XYZ" in html_api)
            
            # Restore original file
            shutil.copy(self.backup_file, self.excel_file)
            os.remove(self.backup_file)
            
            return True
            
        except Exception as e:
            self.log_test("API data modifications", False, str(e))
            if os.path.exists(self.backup_file):
                shutil.copy(self.backup_file, self.excel_file)
                os.remove(self.backup_file)
            return False
    
    def test_3_web_data_modifications(self):
        """Test 3: Verify Web data modifications are reflected in reports"""
        print("\n[TEST 3] Web Data Modifications")
        print("-" * 60)
        
        try:
            # Backup original file
            shutil.copy(self.excel_file, self.backup_file)
            
            from openpyxl import load_workbook
            wb_modify = load_workbook(self.excel_file)
            ws_web = wb_modify['Web Data']
            
            # Modify Test Suite (Row 9 in openpyxl)
            ws_web['A9'] = "MODIFIED_WEB_SUITE_ABC"
            ws_web['B9'] = 555
            ws_web['C9'] = 444
            ws_web['D9'] = 88
            
            # Modify Defect Breakdown (Row 17-18 in openpyxl)
            ws_web['B17'] = 55   # Critical
            ws_web['B18'] = 166  # High
            
            # Modify Coverage (Row 23-24 in openpyxl)
            ws_web['B23'] = 97   # Login&Auth
            ws_web['B24'] = 86   # Checkout
            
            # Modify Risks (Row 31 in openpyxl)
            ws_web['B31'] = "MODIFIED_WEB_RISK_DESCRIPTION_ABC"
            ws_web['D31'] = "MODIFIED_WEB_OWNER_ABC"
            
            wb_modify.save(self.excel_file)
            
            # Generate new report
            html_web, _ = generate_html_report("web")
            
            # --- Assertions ---
            self.log_test("Web Suite Name Change", "MODIFIED_WEB_SUITE_ABC" in html_web)
            self.log_test("Web Total Tests Change", "555" in html_web)
            self.log_test("Web Passed Tests Change", "444" in html_web)
            self.log_test("Web Failed Tests Change", "88" in html_web)
            self.log_test("Web Critical Defects Change", "55" in html_web)
            self.log_test("Web High Defects Change", "166" in html_web)
            self.log_test("Web Login&Auth Coverage Change", "97" in html_web)
            self.log_test("Web Checkout Coverage Change", "86" in html_web)
            self.log_test("Web Risk Description Change", "MODIFIED_WEB_RISK_DESCRIPTION_ABC" in html_web)
            self.log_test("Web Risk Owner Change", "MODIFIED_WEB_OWNER_ABC" in html_web)
            
            # Restore original file
            shutil.copy(self.backup_file, self.excel_file)
            os.remove(self.backup_file)
            
            return True
            
        except Exception as e:
            self.log_test("Web data modifications", False, str(e))
            if os.path.exists(self.backup_file):
                shutil.copy(self.backup_file, self.excel_file)
                os.remove(self.backup_file)
            return False
    
    def run_all_tests(self):
        """Run all tests"""
        print("=" * 60)
        print("QA REPORTING SYSTEM - TEST SUITE")
        print("=" * 60)
        
        results = []
        results.append(self.test_1_report_generation())
        results.append(self.test_2_api_data_modifications())
        results.append(self.test_3_web_data_modifications())
        
        # Summary
        print("\n" + "=" * 60)
        print("TEST SUMMARY")
        print("=" * 60)
        
        total_tests = len(self.test_results)
        passed_tests = sum(1 for r in self.test_results if r['passed'])
        failed_tests = total_tests - passed_tests
        
        print(f"\nTotal Tests: {total_tests}")
        print(f"Passed: {passed_tests}")
        print(f"Failed: {failed_tests}")
        
        if failed_tests == 0:
            print("\n✓ ALL TESTS PASSED!")
        else:
            print(f"\n✗ {failed_tests} TEST(S) FAILED")
            print("\nFailed Tests:")
            for result in self.test_results:
                if not result['passed']:
                    print(f"  - {result['test']}: {result['message']}")
        
        print("=" * 60)
        
        return failed_tests == 0

if __name__ == "__main__":
    suite = TestSuite()
    success = suite.run_all_tests()
    exit(0 if success else 1)
